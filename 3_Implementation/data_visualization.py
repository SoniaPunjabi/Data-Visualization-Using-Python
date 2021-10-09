'''
This file contains all the functions required for data visualization
Author : Sonia
Date : 19th July 2021
Time : 2:00 PM
'''
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart3D, ScatterChart, Series


def barchart(marks_1, marks_2, marks_3, ps_no):
    '''
    Function that generates the overall barchart based on range of percentage
    :param marks_1:Marks obtained in SDLC
    :param marks_2: Marks obtained in Python
    :param marks_3: Marks obtained in MBSE
    :param ps_no: PS numbers of all candidates
    :return: none
    '''
    work_book = Workbook()
    write_data = work_book.active
    grade_1=[0,0,0,0,0,0,0]
    write_data.cell(row=1, column=1).value = "PS NO."
    write_data.cell(row=1, column=2).value = "Percent"
    write_data.cell(row=1, column=3).value = "Students"
    write_data.cell(row=1, column=4).value = "Range"
    write_data.cell(row=1, column=5).value = "Grade"
    for i in range(len(ps_no)):
        overall_percent=(marks_1[i]+marks_2[i]+marks_3[i])*(100/300)
        if overall_percent>90:
            grade_1[0]+=1
        elif overall_percent>=80:
            grade_1[1] += 1
        elif overall_percent>=70:
            grade_1[2] += 1
        elif overall_percent>=60:
            grade_1[3] += 1
        elif overall_percent>=45:
            grade_1[4] += 1
        elif overall_percent>=40:
            grade_1[5] += 1
        elif overall_percent<40:
            grade_1[6] += 1
        write_data.cell(row=i+2,column=1).value=ps_no[i]
        write_data.cell(row=i + 2, column=2).value =overall_percent
    categories=["90-100", "80-90", "70-80", "60-70", "45-60", "40-45", "<40"]
    grade=["S", "A", "B", "C", "D", "E", "F"]
    for i in range(7):
        write_data.cell(row=i+2,column=3).value=grade_1[i]
    for i in range(7):
        write_data.cell(row=i+2,column=4).value=categories[i]
    for i in range(7):
        write_data.cell(row=i+2,column=5).value=grade[i]
    values1 = Reference(write_data, min_col=3, min_row=2, max_col=3, max_row=8)
    # chart = LineChart()
    chart = BarChart3D()
    write_data.add_chart(chart, "G2")
    chart.title = "Overall Bar Chart Based on Percentage"
    chart.y_axis.title = 'Number of Students'
    chart.x_axis.title = 'Range of Percentage'
    chart.add_data(values1)
    series_1 = chart.series[0]
    series_1.marker.symbol = "triangle"
    labels = Reference(write_data, min_col=4, min_row=2, max_row=8)
    chart.set_categories(labels)
    work_book.save("BarChart.xlsx")
    print("Please check the BarChart.xlsl file ")


def pie_chart(marks_1, marks_2, marks_3, ps_no):
    '''
        Function that generates the overall Pie chart based on grades
        :param marks_1:Marks obtained in SDLC
        :param marks_2: Marks obtained in Python
        :param marks_3: Marks obtained in MBSE
        :param ps_no: PS numbers of all candidates
        :return: none
    '''
    work_book = Workbook()
    write_data = work_book.active
    write_data.cell(row=1, column=1).value = "PS NO."
    write_data.cell(row=1, column=2).value = "Percent"
    write_data.cell(row=1, column=3).value = "Students"
    write_data.cell(row=1, column=4).value = "Range"
    write_data.cell(row=1, column=5).value = "Grade"
    grade_1 = [0, 0, 0, 0, 0, 0, 0]
    for i in range(len(ps_no)):
        overall_percent=(marks_1[i]+marks_2[i]+marks_3[i])*(100/300)
        if overall_percent >90:
            grade_1[0] += 1
        elif overall_percent >= 80:
            grade_1[1] += 1
        elif overall_percent >= 70:
            grade_1[2] += 1
        elif overall_percent >= 60:
            grade_1[3] += 1
        elif overall_percent >= 45:
            grade_1[4] += 1
        elif overall_percent >= 40:
            grade_1[5] += 1
        elif overall_percent < 40:
            grade_1[6] += 1

        write_data.cell(row=i + 2, column=1).value = ps_no[i]
        write_data.cell(row=i + 2, column=2).value = overall_percent
    categories = ["90-100", "80-90", "70-80", "60-70", "45-60", "40-45", "<40"]
    grade = ["S", "A", "B", "C", "D", "E", "F"]
    for i in range(7):
        write_data.cell(row=i + 2, column=3).value = grade_1[i]
    for i in range(7):
        write_data.cell(row=i + 2, column=4).value = categories[i]
    for i in range(7):
        write_data.cell(row=i + 2, column=5).value = grade[i]
    chart = PieChart()
    # create data for plotting
    labels = Reference(write_data, min_col=5, min_row=2, max_row=8)
    data = Reference(write_data, min_col=3, min_row=2, max_col=3, max_row=8)
    # adding data to the Pie chart object
    chart.add_data(data)
    # set labels in the chart object
    chart.set_categories(labels)
    # set the title of the chart
    chart.title = "Overall Pie Chart based on Grades "
    # add chart to the sheet
    # the top-left corner of a chart
    # is anchored to cell E2 .
    write_data.add_chart(chart, "G2")
    # save the file
    work_book.save(" PieChart.xlsx")
    print("Please check the PieChart.xlsl file ")


def individual_barchart(marks_1, marks_2, marks_3, ps_no):
    '''
        Function that generates the individual barchart of each candidate based on marks obtained in each subject
        :param marks_1:Marks obtained in SDLC
        :param marks_2: Marks obtained in Python
        :param marks_3: Marks obtained in MBSE
        :param ps_no: PS numbers of all candidates
        :return: none
    '''
    work_book = Workbook()
    write_data = work_book.active
    chart=BarChart3D()
    while True:
        try:
            ps_num= int(input("Enter PS Number of Employee:"))
            if ps_num in ps_no:
                break
            else:
                print("Please enter a valid PS number")

        except ValueError:
            print("Please enter valid PS Number")
    write_data.cell(row=1, column=1).value = "PS NO."
    write_data.cell(row=1, column=2).value = "SDLC"
    write_data.cell(row=1, column=3).value = "PYTHON"
    write_data.cell(row=1, column=4).value = "MBSE."
    if ps_num in ps_no:
        index=ps_no.index(ps_num)
        write_data.cell(row=2, column=1).value =ps_no[index]
        write_data.cell(row=2, column=2).value =marks_1[index]
        write_data.cell(row=2, column=3).value =marks_2[index]
        write_data.cell(row=2, column=4).value =marks_3[index]
        values2 = Reference(write_data, min_col=2, min_row=1, max_col=4, max_row=2)
        write_data.add_chart(chart, "A5")
        chart.title = "Individual Bar chart"
        chart.y_axis.title = 'Marks'
        chart.x_axis.title = 'PS_No:'+str(ps_no[index])
        chart.add_data(data=values2, titles_from_data=True)
        work_book.save("IndividualBarChart.xlsx")
        print("Please check the IndividualBarChart.xlsl file ")
    else:
        print("Oops! PS number not in range, Please try again")

# A custom function to calculate
# probability distribution function


def pdf(sorted_percentt):
    '''
        Function that generates the probability distribution for the bell curve
        :param sorted_percentt: percentages
        :return: probability distribution function
    '''
    mean = np.mean(sorted_percentt)
    std = np.std(sorted_percentt)
    prob_dist = 1 / (std * np.sqrt(2 * np.pi)) * np.exp(- (sorted_percentt - mean) ** 2 / (2 * std ** 2))
    return prob_dist


def bell_curve(marks_1, marks_2, marks_3, ps_no):
    '''
        Function that generates the overall bell curve based on range of percentage
        :param marks_1:Marks obtained in SDLC
        :param marks_2: Marks obtained in Python
        :param marks_3: Marks obtained in MBSE
        :param ps_no: PS numbers of all candidates
        :return: none
    '''
    work_book = Workbook()
    write_data = work_book.active
    grade_1=[0,0,0,0,0,0,0]
    write_data.cell(row=1, column=1).value = "PS NO."
    write_data.cell(row=1, column=2).value = "Percent"
    write_data.cell(row=1, column=3).value = "Students"
    write_data.cell(row=1, column=4).value = "Range"
    write_data.cell(row=1, column=5).value = "Grade"
    write_data.cell(row=1, column=6).value = "Sorted%"
    write_data.cell(row=1, column=7).value = "PDF"
    unsorted_percent=[]
    for i in range(len(ps_no)):
        overall_percent=(marks_1[i]+marks_2[i]+marks_3[i])*(100/300)
        unsorted_percent.append(overall_percent)
        if overall_percent>90:
            grade_1[0]+=1
        elif overall_percent>=80:
            grade_1[1] += 1
        elif overall_percent>=70:
            grade_1[2] += 1
        elif overall_percent>=60:
            grade_1[3] += 1
        elif overall_percent>=45:
            grade_1[4] += 1
        elif overall_percent>=40:
            grade_1[5] += 1
        elif overall_percent<40:
            grade_1[6] += 1
        write_data.cell(row=i+2,column=1).value=ps_no[i]
        write_data.cell(row=i + 2, column=2).value =overall_percent
    categories=["90-100", "80-90", "70-80", "60-70", "45-60", "40-45", "<40"]
    grade=["S", "A", "B", "C", "D", "E", "F"]
    for i in range(7):
        write_data.cell(row=i+2,column=3).value=grade_1[i]
    for i in range(7):
        write_data.cell(row=i+2,column=4).value=categories[i]
    for i in range(7):
        write_data.cell(row=i+2,column=5).value=grade[i]

# To generate an array of sorted_percent-values
    sorted_percent = sorted(unsorted_percent)
    for i in range(len(ps_no)):
        write_data.cell(row=i + 2, column=6).value = sorted_percent[i]

# To generate an array of
# prob_distri-values using corresponding sorted_percent-values
    prob_distri = pdf(sorted_percent)
    for i in range(len(ps_no)):
        write_data.cell(row=i + 2, column=7).value = prob_distri[i]
    chart3 = ScatterChart()
    chart3.title = "Overall Bell curve based on Percentages"
    chart3.x_axis.title = 'Percentages'
    chart3.y_axis.title = 'Probability Distribution'
    chart3.legend = None
    sorted_percent = Reference(write_data, min_col=6, min_row=2, max_row=len(ps_no))
    prob_distri = Reference(write_data, min_col=7, min_row=2, max_row=len(ps_no))
    series_1 = Series(prob_distri, xvalues=sorted_percent)
    chart3.append(series_1)
    write_data.add_chart(chart3, "I2")
    work_book.save("Bell_curve.xlsx")
    print("Please check the Bell_curve.xlsl file ")


def overall_report(marks_1, marks_2, marks_3, ps_no):
    '''
        Function that generates the overall report of all the candidates
        :param marks_1:Marks obtained in SDLC
        :param marks_2: Marks obtained in Python
        :param marks_3: Marks obtained in MBSE
        :param ps_no: PS numbers of all candidates
        :return: none
    '''
    work_book = Workbook()
    write_data = work_book.active
    grade_1=[0,0,0,0,0,0,0]
    write_data.cell(row=1, column=1).value = "PS NO."
    write_data.cell(row=1, column=2).value = "Percent"
    write_data.cell(row=1, column=3).value = "Students"
    write_data.cell(row=1, column=4).value = "Range"
    write_data.cell(row=1, column=5).value = "grade"
    write_data.cell(row=1, column=6).value = "Sorted%"
    write_data.cell(row=1, column=7).value = "PDF"
    unsorted_percent=[]
    for i in range(len(ps_no)):
        overall_percent=(marks_1[i]+marks_2[i]+marks_3[i])*(100/300)
        unsorted_percent.append(overall_percent)
        if overall_percent>90:
            grade_1[0]+=1
        elif overall_percent>=80:
            grade_1[1] += 1
        elif overall_percent>=70:
            grade_1[2] += 1
        elif overall_percent>=60:
            grade_1[3] += 1
        elif overall_percent>=45:
            grade_1[4] += 1
        elif overall_percent>=40:
            grade_1[5] += 1
        elif overall_percent<40:
            grade_1[6] += 1

        write_data.cell(row=i+2,column=1).value=ps_no[i]
        write_data.cell(row=i + 2, column=2).value =overall_percent
    categories=["90-100", "80-90", "70-80", "60-70", "45-60", "40-45", "<40"]
    grade=["S", "A", "B", "C", "D", "E", "F"]
    for i in range(7):
        write_data.cell(row=i+2,column=3).value=grade_1[i]
    for i in range(7):
        write_data.cell(row=i+2,column=4).value=categories[i]
    for i in range(7):
        write_data.cell(row=i+2,column=5).value=grade[i]
# To generate an array of sorted_percent-values
    sorted_percent = sorted(unsorted_percent)
    for i in range(len(ps_no)):
        write_data.cell(row=i + 2, column=6).value = sorted_percent[i]
# To generate an array of
# prob_distri-values using corresponding sorted_percent-values
    prob_distri = pdf(sorted_percent)
    for i in range(len(ps_no)):
        write_data.cell(row=i + 2, column=7).value = prob_distri[i]
    chart3 = ScatterChart()
    chart3.title = "Overall Bell curve based on Percentages"
    chart3.x_axis.title = 'Percentages'
    chart3.y_axis.title = 'Probability Distribution'
    chart3.legend = None
    sorted_percent = Reference(write_data, min_col=6, min_row=2, max_row=len(ps_no))
    prob_distri = Reference(write_data, min_col=7, min_row=2, max_row=len(ps_no))
    series_1 = Series(prob_distri, xvalues=sorted_percent)
    chart3.append(series_1)
    write_data.add_chart(chart3, "I2")
    values1 = Reference(write_data, min_col=3, min_row=2, max_col=3, max_row=8)
    # chart = LineChart()
    chart = BarChart3D()
    write_data.add_chart(chart, "I17")
    chart.title = "Overall Bar Chart Based on Percentage"
    chart.y_axis.title = 'Number of Students'
    chart.x_axis.title = 'Range of Percentage'
    chart.add_data(values1)
    series_1 = chart.series[0]
    series_1.marker.symbol = "triangle"
    labels = Reference(write_data, min_col=4, min_row=2, max_row=8)
    chart.set_categories(labels)
    chart = PieChart()
    # create data for plotting
    labels = Reference(write_data, min_col=5, min_row=2, max_row=8)
    data = Reference(write_data, min_col=3, min_row=2, max_col=3, max_row=8)
    # adding data to the Pie chart object
    chart.add_data(data)
    # set labels in the chart object
    chart.set_categories(labels)
    # set the title of the chart
    chart.title = "Overall Pie Chart based on Grades "
    # add chart to the sheet
    # the top-left corner of a chart
    # is anchored to cell E2 .
    write_data.add_chart(chart, "R2")
    work_book.save("Overall_Report.xlsx")
    print("Please check the Overall_Report.xlsl file ")

